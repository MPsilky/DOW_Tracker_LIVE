# -*- coding: utf-8 -*-
from __future__ import annotations
"""
DOW 30 Tracker — market-hours only (no after-hours)
- Single instance (second run focuses the first)
- Backfill full day on launch; live capture during session
- Buckets: 9:31 AM, 10:00 AM, 11:00 AM, 12 NOON, 1:00 PM, 2:00 PM, 3:00 PM, 4:00 PM
- Arrows / % / colors compare each bucket to the immediately prior bucket (9:31 uses the prior session's 4:00 PM capture, then
  falls back to the latest close if needed)
- Excel export: daily file "Sheet__MM_DD_YYYY.xlsx"
  * Grid sheet mirrors the UI (tickers + all buckets)
  * One sheet per bucket with numeric `price` and `pct` columns
- User-selectable data folder (remembered in settings.json)
"""

import os, sys, json, socket, threading, traceback, re, atexit, io
from pathlib import Path
from datetime import datetime, date
from typing import Any, Dict, Optional, List, Tuple, Callable, Set, cast

try:
    from zoneinfo import ZoneInfo  # type: ignore
except Exception:
    ZoneInfo = None  # type: ignore[assignment]

# -------- optional deps (guarded) --------
try:
    import yfinance as yf   # type: ignore
    import pandas as pd     # type: ignore
    import numpy as np      # type: ignore
except Exception:
    yf = None   # type: ignore[assignment]
    pd = None   # type: ignore[assignment]
    np = None   # type: ignore[assignment]
try:
    import openpyxl  # noqa: F401
    from openpyxl import load_workbook  # type: ignore
except Exception:
    openpyxl = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]

# -------- Qt --------
from PyQt5.QtCore import (
    Qt,
    QTimer,
    QSize,
    QUrl,
    QEasingCurve,
    QPropertyAnimation,
    QSequentialAnimationGroup,
    pyqtProperty,
    pyqtSlot,
    QPoint,
    QPointF,
    QRectF,
)
from PyQt5.QtGui import (
    QIcon,
    QColor,
    QBrush,
    QCloseEvent,
    QFont,
    QDesktopServices,
    QPainter,
    QPen,
    QPaintEvent,
)
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QToolBar, QAction,
    QFileDialog, QMessageBox, QSystemTrayIcon, QMenu, QDialog, QVBoxLayout,
    QHBoxLayout, QLabel, QCheckBox, QPushButton, QStyle, QHeaderView, QTextEdit,
    QWidget, QProgressBar, QFrame, QComboBox, QInputDialog, QSizePolicy
)

APP_NAME = "DOW 30 Tracker"
APP_PORT = 53121  # single-instance IPC

# Order; rows 31/32 are blue
TICKERS: List[str] = [
    "AAPL","AMGN","AXP","BA","CAT","CRM","CSCO","CVX","DIS","DOW","GS","HD",
    "HON","IBM","JNJ","JPM","KO","MCD","MMM","MRK","MSFT","NKE","PG","TRV",
    "UNH","V","VZ","WMT","NVDA","AMZN","INTC","WBA"
]

# Buckets (no After Hours)
BUCKETS: List[Tuple[str, Tuple[int,int]]] = [
    ("9:31 AM", (9,31)),
    ("10:00 AM", (10,0)),
    ("11:00 AM", (11,0)),
    ("12 NOON", (12,0)),
    ("1:00 PM", (13,0)),
    ("2:00 PM", (14,0)),
    ("3:00 PM", (15,0)),
    ("4:00 PM", (16,0)),
]
TIME_COLS: List[str] = [b[0] for b in BUCKETS]
LABEL_TO_INDEX = {label:i for i,(label,_) in enumerate(BUCKETS)}
INDEX_TO_LABEL = {i:label for i,(label,_) in enumerate(BUCKETS)}
MARKET_OPEN = (9,31)
MARKET_CLOSE = (16,0)

# -------- paths & settings --------
def resource_path(rel: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)

def exe_path() -> str:
    return sys.executable if getattr(sys, "frozen", False) else os.path.abspath(sys.argv[0])

if sys.platform.startswith("win"):
    local_appdata = os.environ.get("LOCALAPPDATA")
    if local_appdata:
        STATE_DIR = Path(local_appdata) / "DOW30Tracker"
    else:
        STATE_DIR = Path.home() / "AppData" / "Local" / "DOW30Tracker"
else:
    STATE_DIR = Path.home() / ".dow30tracker"
STATE_DIR.mkdir(parents=True, exist_ok=True)
SETTINGS_JSON = STATE_DIR / "settings.json"
INSTANCE_LOCK = STATE_DIR / "tracker.lock"
_LOCK_HANDLE: Optional["io.TextIOWrapper"] = None  # kept alive for the process lifetime

class _SingleInstanceError(RuntimeError):
    pass

def _acquire_lock() -> None:
    global _LOCK_HANDLE
    if _LOCK_HANDLE is not None:
        return
    try:
        fh = INSTANCE_LOCK.open("a+")
        if os.name == "nt":
            import msvcrt  # type: ignore
            try:
                msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
            except OSError:
                fh.close()
                raise _SingleInstanceError
        else:
            import fcntl  # type: ignore
            try:
                fcntl.lockf(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            except OSError:
                fh.close()
                raise _SingleInstanceError
        _LOCK_HANDLE = fh
    except FileNotFoundError:
        STATE_DIR.mkdir(parents=True, exist_ok=True)
        _acquire_lock()

def _release_lock() -> None:
    global _LOCK_HANDLE
    if _LOCK_HANDLE is None:
        return
    try:
        if os.name == "nt":
            import msvcrt  # type: ignore
            try:
                _LOCK_HANDLE.seek(0)
                msvcrt.locking(_LOCK_HANDLE.fileno(), msvcrt.LK_UNLCK, 1)
            except OSError:
                pass
        else:
            import fcntl  # type: ignore
            try:
                fcntl.lockf(_LOCK_HANDLE.fileno(), fcntl.LOCK_UN)
            except OSError:
                pass
    finally:
        try:
            _LOCK_HANDLE.close()
        except Exception:
            pass
        _LOCK_HANDLE = None

atexit.register(_release_lock)

def _notify_existing_instance() -> None:
    try:
        with socket.create_connection(("127.0.0.1", APP_PORT), timeout=1.5) as c:
            c.sendall(b"SHOW")
    except Exception:
        pass

def ensure_single_instance() -> None:
    try:
        _acquire_lock()
    except _SingleInstanceError:
        _notify_existing_instance()
        sys.exit(0)

def load_settings() -> Dict[str, Any]:
    try:
        if SETTINGS_JSON.exists():
            return json.loads(SETTINGS_JSON.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}

def save_settings(s: Dict[str, Any]) -> None:
    SETTINGS_JSON.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_JSON.write_text(json.dumps(s, indent=2), encoding="utf-8")

def default_data_dir() -> Path:
    doc_candidates: List[Path] = []
    home = Path.home()
    doc_candidates.append(home / "Documents" / "Saved DOW Sheets")
    if sys.platform.startswith("win"):
        # Honour OneDrive documents if present
        for env_key in ("OneDriveCommercial", "OneDriveConsumer", "ONEDRIVE", "OneDrive"):
            onedrive = os.environ.get(env_key)
            if onedrive:
                doc_candidates.insert(0, Path(onedrive) / "Documents" / "Saved DOW Sheets")
    for candidate in doc_candidates:
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
        except Exception:
            continue
    fallback = STATE_DIR / "Saved"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback

def get_data_dir() -> Path:
    s = load_settings()
    raw = s.get("data_dir")
    d = Path(raw) if isinstance(raw, str) and raw else default_data_dir()
    try: d.mkdir(parents=True, exist_ok=True)
    except Exception: d = default_data_dir()
    return d

DATA_DIR = get_data_dir()
LOG_PATH = DATA_DIR / "tracker.log"
FEATURES_JSON = DATA_DIR / "features.json"

_WORKBOOK_RE = re.compile(r"Sheet__(\d{2})_(\d{2})_(\d{4})\.xlsx$", re.IGNORECASE)
_PRIOR_BUCKET_CACHE: Dict[Tuple[str, str], Dict[str, Optional[float]]]
_PRIOR_BUCKET_CACHE = {}


def _parse_workbook_date(path: Path) -> Optional[date]:
    m = _WORKBOOK_RE.match(path.name)
    if not m:
        return None
    try:
        month, day, year = map(int, m.groups())
        return date(year, month, day)
    except ValueError:
        return None

    def _post_capture_hooks(
        self,
        label: str,
        pct_map: Dict[str, Optional[float]],
        capture_dt: Optional[datetime] = None,
    ) -> None:
        self._last_pct_map = pct_map
        self.last_capture_label = label
        if capture_dt is None:
            capture_dt = now_eastern()
        self.last_capture_time = capture_dt
        self._update_dashboard(label, pct_map)
        self._update_right_rail(label, pct_map, capture_dt)
        if self.features.get(FEATURE_NEWS_PING, False):
            self._maybe_news_ping(label, pct_map)
        if self.features.get(FEATURE_HIST_ECHO, False):
            self._maybe_historical_echo(label, pct_map)
        if self.features.get(FEATURE_MARKET_SOUNDS, False):
            self._maybe_play_sounds(label, pct_map)
        if self.features.get(FEATURE_DNA_EXPORT, False):
            self._maybe_export_dna()
        if self.features.get(FEATURE_MORNING_RESUME, False):
            self._maybe_show_morning_resume()
        if self.features.get(FEATURE_SPARKLINE, False):
            self._selection_changed(self.table.currentItem(), None)

    def _maybe_news_ping(self, label: str, pct_map: Dict[str, Optional[float]]) -> None:
        if yf is None:
            return
        if label in self._news_pinged:
            return
        top_ticker: Optional[str] = None
        top_pct: Optional[float] = None
        for tk, pct in pct_map.items():
            if not isinstance(pct, (int, float)):
                continue
            if top_pct is None or abs(pct) > abs(top_pct):
                top_pct = pct
                top_ticker = tk
        if top_ticker is None or top_pct is None or abs(top_pct) < 0.5:
            return

        def worker() -> None:
            try:
                ticker_obj = yf.Ticker(top_ticker)
                news_items = getattr(ticker_obj, "news", None)
                if not news_items:
                    return
                entry = news_items[0]
                headline = entry.get("title") if isinstance(entry, dict) else None
                if not headline:
                    return
                message = f"{top_ticker} {top_pct:+.2f}% — {headline}"

                def show_message() -> None:
                    self.tray.showMessage("Mover Alert", message, QSystemTrayIcon.Information, 6000)

                QTimer.singleShot(0, show_message)
                self._news_pinged[label] = top_ticker
            except Exception as exc:
                log(f"news ping error: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def _daily_history_for(self, tk: str) -> Optional["pd.DataFrame"]:
        if tk in _DAILY_HISTORY_CACHE:
            return _DAILY_HISTORY_CACHE[tk]
        if yf is None or pd is None:
            return None
        try:
            df = yf.Ticker(tk).history(period="1y", interval="1d")
            if _df_empty(df):
                return None
            _DAILY_HISTORY_CACHE[tk] = df
            return df
        except Exception as exc:
            log(f"daily history error {tk}: {exc}")
            return None

    def _maybe_historical_echo(self, label: str, pct_map: Dict[str, Optional[float]]) -> None:
        notified = self._echo_notified.setdefault(label, set())
        for tk, pct in pct_map.items():
            if not isinstance(pct, (int, float)) or abs(pct) < 2.0 or tk in notified:
                continue
            notified.add(tk)

            def worker(ticker: str, pct_val: float) -> None:
                df = self._daily_history_for(ticker)
                if df is None:
                    return
                try:
                    closes = df["Close"].astype(float)
                    returns = closes.pct_change().dropna() * 100.0
                    if returns.empty:
                        return
                    deltas = (returns - pct_val).abs()
                    best_idx = deltas.idxmin()
                    hist_pct = returns.loc[best_idx]
                    msg = f"{ticker} {pct_val:+.2f}% echoes {best_idx.strftime('%b %d, %Y')} ({hist_pct:+.2f}%)."

                    def show() -> None:
                        self.tray.showMessage("Historical Echo", msg, QSystemTrayIcon.Information, 6000)

                    QTimer.singleShot(0, show)
                except Exception as exc:
                    log(f"historical echo error {ticker}: {exc}")

            threading.Thread(target=worker, args=(tk, pct), daemon=True).start()

    def _sound_threshold(self) -> float:
        val = self.settings.get("sound_threshold") if isinstance(self.settings, dict) else None
        try:
            return max(0.1, float(val)) if val is not None else 1.5
        except Exception:
            return 1.5

    def configure_sound_threshold(self) -> None:
        if not self.features.get(FEATURE_MARKET_SOUNDS, False):
            QMessageBox.information(self, APP_NAME, "Enable Custom Market Sounds in Features first.")
            return
        current = self._sound_threshold()
        value, ok = QInputDialog.getDouble(self, APP_NAME, "Play chime when move exceeds (%)", current, 0.1, 10.0, 2)
        if ok:
            self.settings["sound_threshold"] = value
            save_settings(self.settings)
            self.statusBar().showMessage(f"Sound threshold set to {value:.2f}%", 2500)

    def _maybe_play_sounds(self, label: str, pct_map: Dict[str, Optional[float]]) -> None:
        threshold = self._sound_threshold()
        triggered = self._sound_notified.setdefault(label, set())
        for tk, pct in pct_map.items():
            if not isinstance(pct, (int, float)):
                continue
            if abs(pct) >= threshold and tk not in triggered:
                triggered.add(tk)
                QApplication.beep()

    def _maybe_export_dna(self) -> None:
        if yf is None or pd is None:
            return
        today = now_eastern().date()
        if today.weekday() != 4:
            return
        last = self.settings.get("dna_last_export") if isinstance(self.settings, dict) else None
        if last == today.isoformat():
            return

        def worker() -> None:
            close_df = download_close_dataframe(TICKERS, "2mo", "1d")
            if close_df is None or close_df.empty:
                return
            returns = close_df.pct_change().dropna()
            recent = returns.tail(15)
            if recent.empty:
                return
            corr = recent.corr()
            stats = pd.DataFrame({
                "avg_return_%": recent.mean() * 100.0,
                "vol_%": recent.std() * (252 ** 0.5) * 100.0,
            })
            fn = DATA_DIR / f"Dow_DNA__{today.strftime('%Y_%m_%d')}.csv"
            try:
                with open(fn, "w", encoding="utf-8", newline="") as fh:
                    fh.write("# Dow DNA Export generated on " + today.strftime("%Y-%m-%d") + "\n")
                    fh.write("# Correlation Matrix\n")
                    corr.to_csv(fh)
                    fh.write("\n# Summary Stats (% returns, annualized vol)\n")
                    stats.to_csv(fh)
                self.settings["dna_last_export"] = today.isoformat()
                save_settings(self.settings)

                def status() -> None:
                    self.tray.showMessage(APP_NAME, f"Dow DNA export updated:\n{fn.name}", QSystemTrayIcon.Information, 4000)

                QTimer.singleShot(0, status)
            except Exception as exc:
                log(f"dna export error: {exc}")

        threading.Thread(target=worker, daemon=True).start()

    def _maybe_show_morning_resume(self) -> None:
        if not isinstance(self.settings, dict):
            return
        today_key = self.session_date.isoformat()
        if self.settings.get("morning_resume_seen") == today_key:
            return

        def worker() -> None:
            summary = self._build_morning_summary()
            if not summary:
                return

            def show() -> None:
                self.tray.showMessage("Morning Resume", summary, QSystemTrayIcon.Information, 6000)
                self.settings["morning_resume_seen"] = today_key
                save_settings(self.settings)

            QTimer.singleShot(0, show)

        threading.Thread(target=worker, daemon=True).start()

    def _build_morning_summary(self) -> Optional[str]:
        if yf is None or pd is None:
            return None
        tickers = list(TICKERS) + ["^DJI"]
        close_df = download_close_dataframe(tickers, "5d", "1d")
        if close_df is None or close_df.shape[0] < 2:
            return None
        if "^DJI" in close_df.columns:
            index_series = close_df["^DJI"].dropna()
            close_df = close_df.drop(columns=["^DJI"])
        else:
            index_series = None
        latest = close_df.iloc[-1]
        prior = close_df.iloc[-2]
        pct = (latest / prior - 1.0) * 100.0
        leaders = pct.nlargest(3).dropna()
        laggards = pct.nsmallest(3).dropna()
        leader_txt = ", ".join(f"{tk} {val:+.1f}%" for tk, val in leaders.items())
        laggard_txt = ", ".join(f"{tk} {val:+.1f}%" for tk, val in laggards.items())
        index_txt = ""
        if index_series is not None and len(index_series) >= 2:
            idx_pct = (index_series.iloc[-1] / index_series.iloc[-2] - 1.0) * 100.0
            index_txt = f"Dow closed {idx_pct:+.2f}%"
        return f"Leaders: {leader_txt}\nLaggards: {laggard_txt}\n{index_txt}".strip()


def _latest_workbook_before(session_day: date, data_dir: Path) -> Optional[Path]:
    best: Optional[Tuple[date, Path]] = None
    try:
        for path in data_dir.glob("Sheet__*.xlsx"):
            dt = _parse_workbook_date(path)
            if dt is None or dt >= session_day:
                continue
            if best is None or dt > best[0]:
                best = (dt, path)
    except Exception as e:
        log(f"scan workbooks err: {e}")
        return None
    return best[1] if best else None


def _load_bucket_from_workbook(path: Path, bucket_label: str) -> Dict[str, Optional[float]]:
    cache_key = (str(path.resolve()), bucket_label)
    cached = _PRIOR_BUCKET_CACHE.get(cache_key)
    if cached is not None:
        return dict(cached)
    if openpyxl is None:
        return {}
    try:
        from openpyxl import load_workbook

        sheet_name = safe_sheet_name(bucket_label)
        wb = load_workbook(path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[sheet_name]
        result: Dict[str, Optional[float]] = {}
        for row_idx, tk in enumerate(TICKERS, start=2):
            val = ws.cell(row=row_idx, column=2).value
            result[tk] = float(val) if isinstance(val, (int, float)) else None
        wb.close()
        _PRIOR_BUCKET_CACHE[cache_key] = dict(result)
        return result
    except Exception as e:
        log(f"load workbook bucket err: {e}")
        return {}


def prior_session_bucket(session_day: date, bucket_label: str = "4:00 PM", data_dir: Optional[Path] = None) -> Dict[str, Optional[float]]:
    directory = data_dir if data_dir is not None else DATA_DIR
    path = _latest_workbook_before(session_day, directory)
    if path is None:
        return {}
    bucket = _load_bucket_from_workbook(path, bucket_label)
    if not bucket:
        log(f"no prior bucket data for {bucket_label} in {path.name}")
    return bucket


def clear_prior_bucket_cache() -> None:
    _PRIOR_BUCKET_CACHE.clear()

def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"[{ts}] {msg}\n")
    except Exception:
        pass

# -------- IPC: single instance --------
def start_primary_socket() -> socket.socket:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    try:
        s.bind(("127.0.0.1", APP_PORT))
        s.listen(1)
        return s
    except OSError:
        _notify_existing_instance()
        sys.exit(0)

def socket_listener(sock: socket.socket, on_show: Callable[[], None]) -> None:
    def _loop() -> None:
        try:
            while True:
                conn, _ = sock.accept()
                data = conn.recv(32)
                conn.close()
                if data and data.strip() == b"SHOW":
                    on_show()
        except Exception as e:
            log(f"ipc listener error: {e}")
    threading.Thread(target=_loop, daemon=True).start()

# -------- UI helpers --------

class AnimatedToggle(QCheckBox):
    """A drop-in animated toggle that replaces a plain checkbox."""

    _transparent_pen = QPen(Qt.transparent)
    _outline_pen = QPen(QColor("#1f2937"))

    def __init__(
        self,
        parent: Optional[QWidget] = None,
        *,
        bar_color: QColor | Qt.GlobalColor | str = "#1e293b",
        checked_color: str = "#38bdf8",
        handle_color: QColor | Qt.GlobalColor | str = Qt.white,
        pulse_unchecked_color: str = "#33254e7a",
        pulse_checked_color: str = "#6638bdf8",
    ) -> None:
        super().__init__(parent)
        self.setCursor(Qt.PointingHandCursor)
        self.setContentsMargins(10, 0, 10, 0)

        self._bar_brush = QBrush(QColor(bar_color))
        self._bar_checked_brush = QBrush(QColor(checked_color).lighter(120))
        self._handle_brush = QBrush(QColor(handle_color))
        self._handle_checked_brush = QBrush(QColor(checked_color))
        self._pulse_unchecked_brush = QBrush(QColor(pulse_unchecked_color))
        self._pulse_checked_brush = QBrush(QColor(pulse_checked_color))

        self._handle_position = 0.0
        self._pulse_radius = 0.0

        self._slide_anim = QPropertyAnimation(self, b"handle_position", self)
        self._slide_anim.setEasingCurve(QEasingCurve.InOutCubic)
        self._slide_anim.setDuration(180)

        self._pulse_anim = QPropertyAnimation(self, b"pulse_radius", self)
        self._pulse_anim.setDuration(320)
        self._pulse_anim.setStartValue(10.0)
        self._pulse_anim.setEndValue(20.0)

        self._anim_group = QSequentialAnimationGroup(self)
        self._anim_group.addAnimation(self._slide_anim)
        self._anim_group.addAnimation(self._pulse_anim)

        self.stateChanged.connect(self._start_anim)

    def sizeHint(self) -> QSize:  # type: ignore[override]
        return QSize(60, 32)

    def hitButton(self, pos: QPoint) -> bool:  # type: ignore[override]
        return self.contentsRect().contains(pos)

    @pyqtSlot(int)
    def _start_anim(self, value: int) -> None:
        self._anim_group.stop()
        self._slide_anim.setEndValue(1.0 if value else 0.0)
        self._anim_group.start()

    def paintEvent(self, event: QPaintEvent) -> None:  # type: ignore[override]
        rect = self.contentsRect()
        handle_radius = round(0.26 * rect.height())
        handle_radius = max(handle_radius, 12)

        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        bar_rect = QRectF(0.0, 0.0, rect.width() - handle_radius, 0.42 * rect.height())
        bar_rect.moveCenter(rect.center())
        rounding = bar_rect.height() / 2.0

        trail = rect.width() - 2 * handle_radius
        x_pos = rect.x() + handle_radius + trail * self._handle_position

        if self._pulse_anim.state() == QPropertyAnimation.Running:
            painter.setPen(self._transparent_pen)
            painter.setBrush(self._pulse_checked_brush if self.isChecked() else self._pulse_unchecked_brush)
            painter.drawEllipse(QPointF(x_pos, bar_rect.center().y()), self._pulse_radius, self._pulse_radius)

        painter.setPen(self._transparent_pen)
        if self.isChecked():
            painter.setBrush(self._bar_checked_brush)
        else:
            painter.setBrush(self._bar_brush)
        painter.drawRoundedRect(bar_rect, rounding, rounding)

        painter.setPen(self._outline_pen)
        painter.setBrush(self._handle_checked_brush if self.isChecked() else self._handle_brush)
        painter.drawEllipse(QPointF(x_pos, bar_rect.center().y()), handle_radius, handle_radius)
        painter.end()

    @pyqtProperty(float)
    def handle_position(self) -> float:
        return self._handle_position

    @handle_position.setter
    def handle_position(self, value: float) -> None:
        self._handle_position = value
        self.update()

    @pyqtProperty(float)
    def pulse_radius(self) -> float:
        return self._pulse_radius

    @pulse_radius.setter
    def pulse_radius(self, value: float) -> None:
        self._pulse_radius = value
        self.update()

# -------- features (the toggles remain; explainer copy updated) --------
FEATURE_ITEMS: List[Tuple[str,str]] = [
    ("Mini \"Dow Dashboard\" Pulse", "Live banner showing advancers vs decliners, top movers, market mood, and concentration."),
    ("Auto News Ping for Movers", "Tray balloon with the day's biggest mover plus its freshest Yahoo headline."),
    ("Chart Sparkline View", "Unicode sparkline built from 1-minute closes so you can spot slope without leaving the grid."),
    ("Session Insight Rail", "Fills the right rail with capture progress, countdowns, and save-folder status so the layout never feels empty."),
    ("Replay Mode", "Scrub through saved captures to relive the session or clip highlights."),
    ("Morning Resume Summary", "Morning toast recapping yesterday's leaders, laggards, and index finish."),
    ("Historical Echo Mode", "On ±2% swings, surface the closest historical analog and what happened next."),
    ("Dow \"Concentration\" Meter", "Gauge how much of the move the top five components are driving."),
    ("Custom Market Sounds", "Play chimes whenever a component crosses your chosen percentage threshold."),
    ("Candle Ghosts", "Overlay a ghost line at yesterday's close for quick anchoring."),
    ("Daily Confidence Meter", "Animated breadth bar that tracks advancers out of 30 throughout the day."),
    ("\"Dow DNA\" Export", "Weekly CSV drop with correlations, realized vol, and stat blocks."),
]

(
    FEATURE_MINI_DASH,
    FEATURE_NEWS_PING,
    FEATURE_SPARKLINE,
    FEATURE_INSIGHT_RAIL,
    FEATURE_REPLAY,
    FEATURE_MORNING_RESUME,
    FEATURE_HIST_ECHO,
    FEATURE_CONCENTRATION,
    FEATURE_MARKET_SOUNDS,
    FEATURE_CANDLE_GHOSTS,
    FEATURE_CONFIDENCE,
    FEATURE_DNA_EXPORT,
) = [item[0] for item in FEATURE_ITEMS]

# sparkline + feature caches
SPARK_CHARS = "▁▂▃▄▅▆▇█"
_NEWS_CACHE: Dict[str, Any] = {}
_DAILY_HISTORY_CACHE: Dict[str, Optional["pd.DataFrame"]] = {}

# -------- finance helpers and intraday caching --------
def _df_empty(x: Any) -> bool:
    try:
        import pandas as _pd
        return not isinstance(x, _pd.DataFrame) or cast("_pd.DataFrame", x).empty
    except Exception:
        return True

def _ny_tz() -> str:
    return "America/New_York"


def now_eastern() -> datetime:
    if ZoneInfo is not None:
        try:
            return datetime.now(tz=ZoneInfo("America/New_York"))
        except Exception:
            pass
    try:
        import pytz  # type: ignore

        return datetime.now(tz=pytz.timezone("US/Eastern"))
    except Exception:
        return datetime.now()


def format_et(dt: datetime) -> str:
    try:
        text = dt.strftime("%I:%M %p")
        if text.startswith("0"):
            text = text[1:]
        return f"{text} ET"
    except Exception:
        return dt.strftime("%H:%M ET")

# Intraday cache
_intraday_cache: Dict[str, "pd.Series"] = {}

def _load_intraday_cache() -> None:
    """
    Populate `_intraday_cache` with 1-minute close price series for all
    tickers for the current trading day.  A single call to
    `yfinance.download` fetches data efficiently.  If yfinance or
    pandas are unavailable, the function quietly returns.
    """
    global _intraday_cache
    if yf is None or pd is None:
        return
    try:
        df = yf.download(
            tickers=" ".join(TICKERS),
            period="1d",
            interval="1m",
            prepost=False,
            group_by="ticker",
            auto_adjust=False,
            progress=False,
            threads=True,
        )
        new_cache: Dict[str, "pd.Series"] = {}
        for tk in TICKERS:
            s: Optional[pd.Series] = None
            if isinstance(df, dict):
                sub = df.get(tk)
                if sub is not None and "Close" in sub:
                    s = sub["Close"]
            elif tk in getattr(df, "columns", []) or (
                isinstance(getattr(df, "columns", None), pd.MultiIndex)
                and tk in df.columns.get_level_values(0)
            ):
                try:
                    sub_df = df.xs(tk, level=0) if isinstance(df.columns, pd.MultiIndex) else df[tk]  # type: ignore[index]
                    s = sub_df["Close"]  # type: ignore[index]
                except Exception:
                    s = None
            if s is not None and not s.empty:
                if getattr(s.index, "tz", None) is None:
                    s = s.tz_localize(_ny_tz())  # type: ignore[attr-defined]
                else:
                    s = s.tz_convert(_ny_tz())   # type: ignore[attr-defined]
                new_cache[tk] = s.dropna()
        if new_cache:
            _intraday_cache = new_cache
    except Exception as e:
        log(f"_load_intraday_cache err: {e}")

def series_for_day_1m(tk: str) -> Optional["pd.Series"]:
    """
    Return the cached 1-minute close price series for the given ticker
    if available.  If the cache is empty or the ticker has not yet
    been cached, fall back to a fresh yfinance.Ticker() call.  This
    provides an escape hatch if the cache has not been loaded.
    """
    if yf is None or pd is None:
        return None
    s = _intraday_cache.get(tk)
    if s is not None:
        return s
    try:
        df = yf.Ticker(tk).history(period="1d", interval="1m", prepost=False)
        if _df_empty(df):
            return None
        ser = df["Close"]
        if getattr(ser.index, "tz", None) is None:
            ser = ser.tz_localize(_ny_tz())  # type: ignore[attr-defined]
        else:
            ser = ser.tz_convert(_ny_tz())   # type: ignore[attr-defined]
        return ser.dropna()
    except Exception as e:
        log(f"series_for_day_1m({tk}) err: {e}")
        return None

def price_at_or_before_bucket(tk: str, session_day: date, h: int, m: int) -> Optional[float]:
    if pd is None:
        return None
    s = series_for_day_1m(tk)
    if s is None or s.empty:
        return last_close(tk)
    bkt = pd.Timestamp(session_day.year, session_day.month, session_day.day, h, m, tz=_ny_tz())
    try:
        v = s.asof(bkt)  # type: ignore[attr-defined]
        if v is None and len(s) > 0:
            ss = s.loc[:bkt]
            return float(ss.iloc[-1]) if not ss.empty else last_close(tk)
        return float(v) if v is not None else last_close(tk)
    except Exception:
        ss = s.loc[:bkt]
        return float(ss.iloc[-1]) if not ss.empty else last_close(tk)

def yf_hist(tickers: List[str], period: str, interval: str):
    if yf is None: return None
    try:
        return yf.download(
            tickers=" ".join(tickers),
            period=period, interval=interval,
            group_by="ticker", auto_adjust=False, progress=False, threads=True
        )
    except Exception as e:
        log(f"yfinance download err: {e}"); return None


def download_close_dataframe(tickers: List[str], period: str, interval: str) -> Optional["pd.DataFrame"]:
    if yf is None or pd is None:
        return None
    try:
        df = yf.download(
            tickers=" ".join(tickers),
            period=period,
            interval=interval,
            group_by="ticker",
            auto_adjust=False,
            progress=False,
            threads=True,
        )
        if isinstance(df, dict):
            frames: Dict[str, "pd.Series"] = {}
            for tk in tickers:
                sub = df.get(tk)
                if sub is not None and "Close" in sub:
                    frames[tk] = sub["Close"]
            if not frames:
                return None
            return pd.DataFrame(frames).dropna(how="all")
        if isinstance(getattr(df, "columns", None), pd.MultiIndex):
            try:
                close_df = df.xs("Close", level=1, axis=1)
            except Exception:
                close_df = df
        else:
            close_df = df.get("Close") if hasattr(df, "get") else None
        if close_df is None:
            return None
        if isinstance(close_df, pd.Series):
            close_df = close_df.to_frame(name=tickers[0])
        return cast("pd.DataFrame", close_df).dropna(how="all")
    except Exception as e:
        log(f"download_close_dataframe err: {e}")
        return None

def last_close(tk: str) -> Optional[float]:
    if yf is None: return None
    try:
        h: Any = yf.Ticker(tk).history(period="5d", interval="1d")
        if _df_empty(h): return None
        return float(cast("pd.DataFrame", h)["Close"].iloc[-1])
    except Exception as e:
        log(f"last_close({tk}) err: {e}"); return None

# -------- UI helpers --------
GREEN = QColor("#22c55e")
RED   = QColor("#f87171")
NEUT  = QColor("#e2e8f0")

def make_cell(text: str, pct: Optional[float]) -> QTableWidgetItem:
    it = QTableWidgetItem(text)
    col = GREEN if (isinstance(pct,(int,float)) and pct>0) else RED if (isinstance(pct,(int,float)) and pct<0) else NEUT
    it.setForeground(QBrush(col))
    it.setFlags(it.flags() & ~Qt.ItemIsEditable)
    return it

def safe_sheet_name(label: str) -> str:
    s = label.replace(":", " ").replace("\\","_").replace("/","_").replace("?","_").replace("*","_").replace("[","(").replace("]",")")
    s = s.strip() or "Capture"
    return s[:31]

def day_file_name(dt: datetime) -> str:
    return f"Sheet__{dt.strftime('%m_%d_%Y')}.xlsx"

_price_re = re.compile(r"[-+]?\d+(?:\.\d+)?")
_pct_re   = re.compile(r"\(([+-]?\d+(?:\.\d+)?)%")

def parse_price_pct(cell_text: str) -> Tuple[Optional[float], Optional[float]]:
    s = cell_text.strip()
    s = s.lstrip("▲▼• ").replace("\xa0", " ")
    m1 = _price_re.search(s)
    px = float(m1.group(0)) if m1 else None
    m2 = _pct_re.search(s)
    pct = float(m2.group(1)) if m2 else None
    return px, pct


class DashboardWidget(QFrame):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("DashboardWidget")
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            "QFrame#DashboardWidget {"
            " background-color: #13203a;"
            " border: 1px solid #1f2a40;"
            " border-radius: 10px;"
            " padding: 12px;"
            " }"
        )

        layout = QHBoxLayout(self)
        layout.setContentsMargins(14, 6, 14, 6)
        layout.setSpacing(22)

        font_big = QFont(); font_big.setPointSize(14); font_big.setBold(True)
        font_small = QFont(); font_small.setPointSize(11)

        self.mood_label = QLabel("Mood: --")
        self.mood_label.setFont(font_big)
        self.mood_label.setStyleSheet("color: #f8fafc;")
        layout.addWidget(self.mood_label)

        self.adv_decl_label = QLabel("Advancers 0 · Decliners 0")
        self.adv_decl_label.setFont(font_small)
        layout.addWidget(self.adv_decl_label)

        self.top_mover_label = QLabel("Top mover: --")
        self.top_mover_label.setFont(font_small)
        layout.addWidget(self.top_mover_label)

        self.concentration_label = QLabel("Top 5 share: --")
        self.concentration_label.setFont(font_small)
        layout.addWidget(self.concentration_label)

        self.confidence_bar = QProgressBar()
        self.confidence_bar.setRange(0, len(TICKERS))
        self.confidence_bar.setValue(0)
        self.confidence_bar.setTextVisible(True)
        self.confidence_bar.setFormat("Confidence %v/30")
        self.confidence_bar.setStyleSheet(
            "QProgressBar { background-color: #0b1220; border-radius: 6px;"
            " color: #f8fafc; font-weight: 600; padding: 2px 6px; }"
            "QProgressBar::chunk { background-color: #22c55e; border-radius: 6px; }"
        )
        layout.addWidget(self.confidence_bar)

        layout.addStretch(1)

    def update_metrics(
        self,
        adv: int,
        dec: int,
        flat: int,
        top_ticker: Optional[str],
        top_pct: Optional[float],
        mood: str,
        mood_color: str,
        concentration: Optional[float],
    ) -> None:
        self.adv_decl_label.setText(f"Advancers {adv} · Decliners {dec} · Flat {flat}")
        top_text = "--" if top_ticker is None or top_pct is None else f"{top_ticker} {top_pct:+.2f}%"
        self.top_mover_label.setText(f"Top mover: {top_text}")
        if isinstance(concentration, (int, float)):
            self.concentration_label.setText(f"Top 5 share: {concentration:.1f}%")
        else:
            self.concentration_label.setText("Top 5 share: --")
        self.mood_label.setText(f"Mood: {mood}")
        self.mood_label.setStyleSheet(f"color: {mood_color};")
        self.confidence_bar.setValue(adv)


def _sparkline_for_series(vals: List[float]) -> str:
    if not vals:
        return ""
    lo = min(vals)
    hi = max(vals)
    if hi - lo < 1e-9:
        return SPARK_CHARS[0] * len(vals)
    rng = hi - lo
    chars = []
    for v in vals:
        idx = int((v - lo) / rng * (len(SPARK_CHARS) - 1))
        chars.append(SPARK_CHARS[idx])
    return "".join(chars)


class SparklinePane(QFrame):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("SparklinePane")
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            "QFrame#SparklinePane {"
            " background-color: #101b33;"
            " border: 1px solid #1f2a40;"
            " border-radius: 10px;"
            " padding: 12px;"
            " }"
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(8)

        title_font = QFont(); title_font.setPointSize(13); title_font.setBold(True)
        self.title_label = QLabel("Sparkline")
        self.title_label.setFont(title_font)
        layout.addWidget(self.title_label)

        mono_font = QFont("Consolas")
        mono_font.setPointSize(18)
        self.sparkline_label = QLabel("Select a ticker to preview the slope")
        self.sparkline_label.setFont(mono_font)
        self.sparkline_label.setStyleSheet("color: #38bdf8;")
        self.sparkline_label.setWordWrap(True)
        layout.addWidget(self.sparkline_label)

        self.caption_label = QLabel("")
        self.caption_label.setWordWrap(True)
        layout.addWidget(self.caption_label)

        layout.addStretch(1)

    def show_series(
        self,
        ticker: str,
        series: Optional["pd.Series"],
        latest_price: Optional[float],
    ) -> None:
        if series is None or series.empty:
            self.title_label.setText(f"{ticker} — sparkline")
            self.sparkline_label.setText("No minute data yet today.")
            self.caption_label.setText("")
            return
        vals = series.dropna().tolist()[-64:]
        spark = _sparkline_for_series(vals)
        self.title_label.setText(f"{ticker} — {len(vals)}m trail")
        self.sparkline_label.setText(spark)
        last_ts = series.index[-1]
        ts_text = getattr(last_ts, "strftime", lambda _: "")("%I:%M %p").lstrip("0")
        px_txt = "--" if latest_price is None else f"{latest_price:.2f}"
        self.caption_label.setText(f"Last @ {ts_text} ET — {px_txt}")


class InsightPanel(QFrame):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("InsightPanel")
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet(
            "QFrame#InsightPanel {"
            " background-color: #0e1627;"
            " border: 1px solid #1f2a40;"
            " border-radius: 10px;"
            " padding: 14px;"
            " }"
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(8)

        title_font = QFont(); title_font.setPointSize(13); title_font.setBold(True)
        body_font = QFont(); body_font.setPointSize(12)

        self.title_label = QLabel("Session Insights")
        self.title_label.setFont(title_font)
        layout.addWidget(self.title_label)

        self.session_label = QLabel("Session: --")
        self.session_label.setFont(body_font)
        layout.addWidget(self.session_label)

        self.last_capture_label = QLabel("Last capture: --")
        self.last_capture_label.setFont(body_font)
        layout.addWidget(self.last_capture_label)

        self.breadth_label = QLabel("Breadth: --")
        self.breadth_label.setFont(body_font)
        layout.addWidget(self.breadth_label)

        self.top_label = QLabel("Top mover: --")
        self.top_label.setFont(body_font)
        layout.addWidget(self.top_label)

        self.next_label = QLabel("Next bucket: --")
        self.next_label.setFont(body_font)
        layout.addWidget(self.next_label)

        self.progress = QProgressBar()
        self.progress.setRange(0, len(BUCKETS))
        self.progress.setValue(0)
        self.progress.setTextVisible(True)
        self.progress.setFormat("0/0 buckets")
        self.progress.setStyleSheet(
            "QProgressBar { background-color: #0b1220; border-radius: 6px;"
            " color: #f8fafc; font-weight: 600; padding: 2px 6px; }"
            "QProgressBar::chunk { background-color: #38bdf8; border-radius: 6px; }"
        )
        layout.addWidget(self.progress)

        self.folder_label = QLabel("Saving to: --")
        self.folder_label.setFont(body_font)
        self.folder_label.setWordWrap(True)
        self.folder_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(self.folder_label)

        layout.addStretch(1)

    def update_state(
        self,
        session_text: str,
        last_capture_text: str,
        breadth_text: str,
        top_text: str,
        next_text: str,
        progress_value: int,
        progress_total: int,
        data_dir: Path,
    ) -> None:
        self.session_label.setText(session_text)
        self.last_capture_label.setText(last_capture_text)
        self.breadth_label.setText(breadth_text)
        self.top_label.setText(top_text)
        self.next_label.setText(next_text)
        self.progress.setRange(0, max(progress_total, 1))
        self.progress.setValue(max(0, min(progress_value, progress_total)))
        self.progress.setFormat(f"{progress_value}/{progress_total} buckets")
        self.folder_label.setText(f"Saving to: {data_dir}")


class RightRailWidget(QFrame):
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setObjectName("RightRailWidget")
        self.setFrameShape(QFrame.NoFrame)
        self.setStyleSheet(
            "QFrame#RightRailWidget {"
            " background-color: transparent;"
            " }"
        )

        self.sparkline = SparklinePane()
        self.insights = InsightPanel()

        self.sparkline.setMinimumWidth(260)
        self.insights.setMinimumWidth(260)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)
        layout.addWidget(self.sparkline)
        layout.addWidget(self.insights)
        layout.addStretch(1)

        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)

    def set_modes(self, spark_on: bool, insight_on: bool) -> None:
        self.sparkline.setVisible(spark_on)
        self.insights.setVisible(insight_on)
        self.setVisible(spark_on or insight_on)

    def update_insights(
        self,
        session_text: str,
        last_capture_text: str,
        breadth_text: str,
        top_text: str,
        next_text: str,
        progress_value: int,
        progress_total: int,
        data_dir: Path,
    ) -> None:
        self.insights.update_state(
            session_text,
            last_capture_text,
            breadth_text,
            top_text,
            next_text,
            progress_value,
            progress_total,
            data_dir,
        )


class ReplayDialog(QDialog):
    def __init__(self, parent: QMainWindow, workbook: Path):
        super().__init__(parent)
        self.setWindowTitle("Session Replay")
        self.setWindowIcon(parent.windowIcon())
        self.resize(900, 640)

        self.workbook = workbook
        self.bucket_payload: Dict[str, List[Tuple[str, Optional[float], Optional[float]]]] = {}

        layout = QVBoxLayout(self)
        top_row = QHBoxLayout()
        layout.addLayout(top_row)

        self.bucket_combo = QComboBox()
        for label in TIME_COLS:
            self.bucket_combo.addItem(label)
        top_row.addWidget(QLabel(workbook.name))
        top_row.addStretch(1)
        top_row.addWidget(QLabel("Jump to bucket:"))
        top_row.addWidget(self.bucket_combo)

        self.summary_label = QLabel("")
        layout.addWidget(self.summary_label)

        self.table = QTableWidget(len(TICKERS), 3)
        self.table.setHorizontalHeaderLabels(["Ticker", "Price", "% vs prior"])
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)

        self.bucket_combo.currentTextChanged.connect(self._render_bucket)
        self._load_workbook()
        self._render_bucket(self.bucket_combo.currentText())

    def _load_workbook(self) -> None:
        if openpyxl is None:
            QMessageBox.warning(self, APP_NAME, "Replay requires openpyxl installed.")
            return
        try:
            wb = load_workbook(self.workbook, data_only=True, read_only=True)
        except Exception as e:
            QMessageBox.warning(self, APP_NAME, f"Unable to open workbook:\n{e}")
            return
        try:
            for label in TIME_COLS:
                sheet_name = safe_sheet_name(label)
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                rows: List[Tuple[str, Optional[float], Optional[float]]] = []
                for idx, tk in enumerate(TICKERS, start=2):
                    px = ws.cell(row=idx, column=2).value
                    pct = ws.cell(row=idx, column=3).value
                    rows.append((tk, float(px) if isinstance(px, (int, float)) else None,
                                 float(pct) if isinstance(pct, (int, float)) else None))
                self.bucket_payload[label] = rows
        finally:
            wb.close()

    def _render_bucket(self, label: str) -> None:
        rows = self.bucket_payload.get(label, [])
        adv = dec = 0
        top_ticker = None
        top_pct: Optional[float] = None
        for ridx, (tk, px, pct) in enumerate(rows):
            it_ticker = QTableWidgetItem(f"{ridx+1}. {tk}")
            it_ticker.setFlags(it_ticker.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(ridx, 0, it_ticker)

            txt_px = "--" if px is None else f"{px:.2f}"
            it_px = QTableWidgetItem(txt_px)
            it_px.setFlags(it_px.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(ridx, 1, it_px)

            txt_pct = "--" if pct is None else f"{pct:+.2f}%"
            it_pct = QTableWidgetItem(txt_pct)
            it_pct.setFlags(it_pct.flags() & ~Qt.ItemIsEditable)
            if isinstance(pct, (int, float)):
                if pct > 0:
                    adv += 1
                    it_pct.setForeground(QBrush(GREEN))
                elif pct < 0:
                    dec += 1
                    it_pct.setForeground(QBrush(RED))
                else:
                    it_pct.setForeground(QBrush(NEUT))
                if top_pct is None or abs(pct) > abs(top_pct):
                    top_pct = pct
                    top_ticker = tk
            self.table.setItem(ridx, 2, it_pct)

        flats = len(rows) - adv - dec
        summary = f"Advancers {adv} · Decliners {dec} · Flat {flats}"
        if isinstance(top_pct, (int, float)) and top_ticker:
            summary += f" — Top mover {top_ticker} {top_pct:+.2f}%"
        self.summary_label.setText(summary)
# -------- Main Window --------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_NAME); self.setWindowIcon(self._icon()); self.resize(1400, 860)

        self.settings = load_settings()
        self.features = self._load_features()
        self.session_date = now_eastern().date()
        self.prev_close: Dict[str, Optional[float]] = {}
        self.prev_session_bucket: Dict[str, Optional[float]] = prior_session_bucket(self.session_date)
        self.bucket_prices: Dict[str, Dict[str, Optional[float]]] = {tk: {} for tk in TICKERS}
        self._exported_states: Dict[str, Tuple[Optional[float], ...]] = {}
        self._last_timer_key: Optional[Tuple[int, int]] = None
        self._news_pinged: Dict[str, str] = {}
        self._echo_notified: Dict[str, Set[str]] = {}
        self._sound_notified: Dict[str, Set[str]] = {}
        self._last_pct_map: Dict[str, Optional[float]] = {}
        self.last_capture_label: Optional[str] = None
        self.last_capture_time: Optional[datetime] = None

        # table
        self.table = QTableWidget(len(TICKERS), 1 + len(TIME_COLS))
        self.table.setHorizontalHeaderLabels(["Ticker"] + TIME_COLS)
        self.table.verticalHeader().setVisible(False)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        hh.setDefaultAlignment(Qt.AlignCenter)
        hh.setSectionResizeMode(0, QHeaderView.Fixed)

        base_font = QFont(); base_font.setPointSize(13)
        self.setFont(base_font)
        self.table.setFont(base_font)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setAlternatingRowColors(True)
        for r in range(len(TICKERS)):
            self.table.setRowHeight(r, 34)

        fbold = QFont(); fbold.setPointSize(13); fbold.setBold(True)
        first_col_bg = QBrush(QColor("#1e293b"))
        for i, tk in enumerate(TICKERS, start=1):
            it = QTableWidgetItem(f"{i}. {tk}")
            it.setFont(fbold)
            color = QColor("#38bdf8") if tk in ("INTC","WBA") else QColor("#f8fafc")
            it.setForeground(QBrush(color))
            it.setBackground(first_col_bg)
            it.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)
            it.setFlags(it.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(i-1, 0, it)
        self.table.setColumnWidth(0, 150)

        self.dashboard = DashboardWidget()
        self.right_rail = RightRailWidget()
        self.sparkline = self.right_rail.sparkline

        central = QWidget()
        central_layout = QVBoxLayout(central)
        central_layout.setContentsMargins(14, 10, 14, 10)
        central_layout.setSpacing(10)
        central_layout.addWidget(self.dashboard)

        main_row = QHBoxLayout()
        main_row.setSpacing(10)
        main_row.addWidget(self.table, 1)
        main_row.addWidget(self.right_rail)
        central_layout.addLayout(main_row, 1)
        self.setCentralWidget(central)

        # toolbar
        tb = QToolBar("Main"); tb.setIconSize(QSize(22,22)); tb.setMovable(False)
        ftb = QFont(); ftb.setPointSize(12); ftb.setBold(True); tb.setFont(ftb)
        tb.setToolButtonStyle(Qt.ToolButtonTextOnly)
        self.addToolBar(Qt.TopToolBarArea, tb)

        self.setStyleSheet(
            "QMainWindow { background-color: #0f172a; }\n"
            "QToolBar { background: transparent; padding: 12px; spacing: 6px; }\n"
            "QToolButton { color: #f8fafc; padding: 8px 14px; border-radius: 6px; background-color: rgba(37, 99, 235, 0.18); }\n"
            "QToolButton:hover { background-color: #2563eb; color: #f8fafc; }\n"
            "QToolButton:checked { background-color: #1d4ed8; color: #f8fafc; }\n"
            "QTableWidget { background-color: #0b1220; color: #e2e8f0; gridline-color: #1f2a40; alternate-background-color: #131d33; selection-background-color: #2563eb; selection-color: #f8fafc; border: 1px solid #1f2a40; }\n"
            "QTableWidget::item { padding: 6px 10px; }\n"
            "QHeaderView::section { background-color: #1e293b; color: #f8fafc; font-size: 13px; padding: 10px 12px; border: 0; }\n"
            "QStatusBar { background: #111827; color: #f8fafc; font-weight: 600; padding: 4px 12px; }\n"
            "QMenu { background: #0f172a; color: #f8fafc; border: 1px solid #1f2a40; }\n"
            "QMenu::item:selected { background: #2563eb; }\n"
            "QDialog { background-color: #0f172a; color: #f1f5f9; }\n"
            "QLabel { color: #f1f5f9; }\n"
            "QTextEdit { background-color: #0b1220; color: #e2e8f0; border: 1px solid #1f2a40; }\n"
            "QPushButton { background-color: #1d4ed8; border-radius: 6px; padding: 8px 16px; color: #f8fafc; font-weight: 600; }\n"
            "QPushButton:hover { background-color: #2563eb; }\n"
            "QCheckBox { color: #f1f5f9; font-size: 13px; }"
        )

        self.actRefresh = QAction("Refresh", self); self.actRefresh.triggered.connect(self.refresh_now); tb.addAction(self.actRefresh)
        self.actBrowse  = QAction("Browse Excels", self); self.actBrowse.triggered.connect(self.browse_excels); tb.addAction(self.actBrowse)

        tb.addSeparator()
        self.actAutoOn  = QAction("Enable Auto-Start on Login", self);  self.actAutoOn.triggered.connect(self.enable_autostart); tb.addAction(self.actAutoOn)
        self.actAutoOff = QAction("Disable Auto-Start on Login", self); self.actAutoOff.triggered.connect(self.disable_autostart); tb.addAction(self.actAutoOff)

        tb.addSeparator()
        self.optTimes = QAction("Show Times", self, checkable=True, checked=True)
        self.optPct   = QAction("Show %",     self, checkable=True, checked=True)
        self.optArw   = QAction("Arrows",     self, checkable=True, checked=True)
        self.optStripe= QAction("Stripe Rows",self, checkable=True, checked=True)
        for a in (self.optTimes, self.optPct, self.optArw, self.optStripe): tb.addAction(a)
        self.optTimes.toggled.connect(self._toggle_times)
        self.optPct.toggled.connect(lambda _: self._rerender_table())
        self.optArw.toggled.connect(lambda _: self._rerender_table())
        self.optStripe.toggled.connect(self._toggle_stripes)

        tb.addSeparator()
        self.actFeatures = QAction("Features", self); self.actFeatures.triggered.connect(self.open_features); tb.addAction(self.actFeatures)
        self.actGuide    = QAction("Where the hell is the thing?", self); self.actGuide.triggered.connect(self.show_guide); tb.addAction(self.actGuide)

        self.actReplay = QAction("Replay Session", self); self.actReplay.triggered.connect(self.open_replay); tb.addAction(self.actReplay)
        self.actSoundThreshold = QAction("Set Sound Threshold…", self); self.actSoundThreshold.triggered.connect(self.configure_sound_threshold); tb.addAction(self.actSoundThreshold)

        tb.addSeparator()
        self.actSetFolder = QAction("Set Data Folder…", self); self.actSetFolder.triggered.connect(self.set_data_folder); tb.addAction(self.actSetFolder)

        # tray
        self.tray = QSystemTrayIcon(self._icon(), self)
        menu = QMenu()
        menu.addAction("Show").triggered.connect(self.show_main)
        menu.addAction("Exit").triggered.connect(self.exit_app)
        self.tray.setContextMenu(menu); self.tray.setToolTip(APP_NAME); self.tray.show()
        self.tray.activated.connect(
            lambda reason: self.show_main() if reason in (QSystemTrayIcon.Trigger, QSystemTrayIcon.DoubleClick) else None
        )

        app = QApplication.instance()
        if app: app.setQuitOnLastWindowClosed(False)

        self.live_timer = QTimer(self); self.live_timer.timeout.connect(self.timer_logic); self.live_timer.start(60_000)
        QTimer.singleShot(60, self.initial_sync)

        status_font = QFont(); status_font.setPointSize(11); status_font.setBold(True)
        self.statusBar().setFont(status_font)
        self.statusBar().setSizeGripEnabled(False)
        self.statusBar().showMessage("Ready — waiting for the first capture at 9:31 AM ET.")

        self.table.currentItemChanged.connect(self._selection_changed)
        self.right_rail.update_insights(
            f"Session: {self.session_date.strftime('%A, %b %d, %Y')}",
            "Last capture: --",
            "Breadth: --",
            "Top mover: --",
            "Next bucket: --",
            0,
            len(BUCKETS),
            DATA_DIR,
        )
        self._apply_feature_visibility()

    # ---------- core rendering ----------
    def _reset_for_new_session(self, new_date: Optional[date] = None) -> None:
        self.session_date = new_date or now_eastern().date()
        self.prev_close = {}
        clear_prior_bucket_cache()
        self.prev_session_bucket = prior_session_bucket(self.session_date)
        self.bucket_prices = {tk: {} for tk in TICKERS}
        self._exported_states = {}
        self._last_timer_key = None
        self._news_pinged = {}
        self._echo_notified = {}
        self._sound_notified = {}
        self._last_pct_map = {}
        self.last_capture_label = None
        self.last_capture_time = None
        for r in range(len(TICKERS)):
            for c in range(1, 1 + len(TIME_COLS)):
                blank = QTableWidgetItem("")
                blank.setFlags(blank.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(r, c, blank)
        self._update_dashboard()
        self._update_right_rail()

    def _ensure_session(self, current_date: Optional[date] = None) -> None:
        target = current_date or now_eastern().date()
        if target != self.session_date:
            self._reset_for_new_session(target)

    def _ensure_prev_close(self, tk: str) -> Optional[float]:
        if tk not in self.prev_close or self.prev_close[tk] is None:
            self.prev_close[tk] = last_close(tk)
        return self.prev_close.get(tk)

    def _base_for(self, tk: str, label: str) -> Optional[float]:
        idx = LABEL_TO_INDEX[label]
        if idx == 0:
            prior_val = self.prev_session_bucket.get(tk)
            if isinstance(prior_val, (int, float)):
                return prior_val
            return self._ensure_prev_close(tk)

        bucket_dict = self.bucket_prices.get(tk, {})
        for prev_idx in range(idx - 1, -1, -1):
            prev_label = INDEX_TO_LABEL[prev_idx]
            prev_val = bucket_dict.get(prev_label)
            if isinstance(prev_val, (int, float)):
                return prev_val

        return self._ensure_prev_close(tk)

    def _render_cell(self, tk: str, label: str, price: Optional[float], base: Optional[float]) -> Tuple[str, Optional[float], QTableWidgetItem]:
        pct: Optional[float] = None
        if isinstance(price, (int, float)) and isinstance(base, (int, float)) and base:
            pct = ((price / base) - 1.0) * 100.0

        text = "--" if not isinstance(price, (int, float)) else f"{price:.2f}"
        if self.optPct.isChecked() and isinstance(pct, (int, float)):
            text += f"  ({pct:+.2f}%)"
        if self.optArw.isChecked() and isinstance(pct, (int, float)):
            text = ("▲ " if pct > 0 else ("▼ " if pct < 0 else "• ")) + text

        it = QTableWidgetItem(text)
        col = GREEN if (isinstance(pct,(int,float)) and pct>0) else RED if (isinstance(pct,(int,float)) and pct<0) else NEUT
        it.setForeground(QBrush(col))
        if self.features.get(FEATURE_CANDLE_GHOSTS, False):
            prev = self._ensure_prev_close(tk)
            bg_color = QColor("#1e293b")
            if isinstance(prev, (int, float)) and isinstance(price, (int, float)):
                if price > prev:
                    bg_color = QColor("#0f3d2e")
                elif price < prev:
                    bg_color = QColor("#3b1020")
                else:
                    bg_color = QColor("#1e293b")
            it.setBackground(QBrush(bg_color))
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
        return text, pct, it

    def _rerender_table(self) -> None:
        for label in TIME_COLS:
            col = 1 + LABEL_TO_INDEX[label]
            for row, tk in enumerate(TICKERS):
                price = self.bucket_prices.get(tk, {}).get(label)
                base = self._base_for(tk, label)
                _txt, pct, item = self._render_cell(tk, label, price, base)
                self.table.setItem(row, col, item)
        self._update_dashboard()

    def _toggle_times(self, checked: bool) -> None:
        self.table.horizontalHeader().setVisible(checked)

    def _toggle_stripes(self, checked: bool) -> None:
        self.table.setAlternatingRowColors(checked)
        self.table.viewport().update()

    def _apply_feature_visibility(self) -> None:
        show_dash = any(self.features.get(key, False) for key in (FEATURE_MINI_DASH, FEATURE_CONCENTRATION, FEATURE_CONFIDENCE))
        self.dashboard.setVisible(show_dash)
        spark_on = self.features.get(FEATURE_SPARKLINE, False)
        insight_on = self.features.get(FEATURE_INSIGHT_RAIL, True)
        self.right_rail.set_modes(spark_on, insight_on)
        self.actReplay.setVisible(self.features.get(FEATURE_REPLAY, False))
        self.actSoundThreshold.setVisible(self.features.get(FEATURE_MARKET_SOUNDS, False))
        if spark_on:
            QTimer.singleShot(0, lambda: self._selection_changed(self.table.currentItem(), None))
        else:
            self.sparkline.show_series("Ticker", None, None)
        if show_dash:
            self._update_dashboard()
        self._update_right_rail()

    def _selection_changed(self, current: Optional[QTableWidgetItem], previous: Optional[QTableWidgetItem]) -> None:
        if not self.features.get(FEATURE_SPARKLINE, False):
            return
        row = self.table.currentRow()
        if row < 0 or row >= len(TICKERS):
            return
        tk = TICKERS[row]
        series = series_for_day_1m(tk)
        label = self._latest_label_with_data()
        latest_price = None
        if label is not None:
            latest_price = self.bucket_prices.get(tk, {}).get(label)
        self.sparkline.show_series(tk, series, latest_price)

    def _latest_label_with_data(self) -> Optional[str]:
        for label in reversed(TIME_COLS):
            for tk in TICKERS:
                if isinstance(self.bucket_prices.get(tk, {}).get(label), (int, float)):
                    return label
        return None

    def _update_dashboard(self, label: Optional[str] = None, pct_map: Optional[Dict[str, Optional[float]]] = None) -> None:
        if not any(self.features.get(key, False) for key in (FEATURE_MINI_DASH, FEATURE_CONCENTRATION, FEATURE_CONFIDENCE)):
            return
        label = label or self._latest_label_with_data()
        if label is None:
            self.dashboard.update_metrics(0, 0, len(TICKERS), None, None, "Calm", "#f8fafc", None)
            return
        if pct_map is None:
            pct_map = {}
            for tk in TICKERS:
                price = self.bucket_prices.get(tk, {}).get(label)
                base = self._base_for(tk, label)
                pct_val: Optional[float] = None
                if isinstance(price, (int, float)) and isinstance(base, (int, float)) and base:
                    pct_val = ((price / base) - 1.0) * 100.0
                pct_map[tk] = pct_val

        adv = sum(1 for pct in pct_map.values() if isinstance(pct, (int, float)) and pct > 0)
        dec = sum(1 for pct in pct_map.values() if isinstance(pct, (int, float)) and pct < 0)
        flat = len(TICKERS) - adv - dec
        top_ticker: Optional[str] = None
        top_pct: Optional[float] = None
        for tk, pct in pct_map.items():
            if not isinstance(pct, (int, float)):
                continue
            if top_pct is None or abs(pct) > abs(top_pct):
                top_pct = pct
                top_ticker = tk

        mood = "Balanced"
        mood_color = "#facc15"
        if adv - dec >= 10:
            mood = "Bullish"
            mood_color = "#22c55e"
        elif dec - adv >= 10:
            mood = "Bearish"
            mood_color = "#f87171"

        pct_values = [abs(pct) for pct in pct_map.values() if isinstance(pct, (int, float))]
        concentration = None
        if pct_values:
            total = sum(pct_values)
            if total:
                top_five = sorted(pct_values, reverse=True)[:5]
                concentration = (sum(top_five) / total) * 100.0

        self.dashboard.update_metrics(adv, dec, flat, top_ticker, top_pct, mood, mood_color, concentration)

    def _captured_bucket_count(self) -> int:
        count = 0
        for label in TIME_COLS:
            for tk in TICKERS:
                if isinstance(self.bucket_prices.get(tk, {}).get(label), (int, float)):
                    count += 1
                    break
        return count

    def _next_bucket_context(self) -> Tuple[Optional[str], str]:
        now_dt = now_eastern()
        for label, (h, m) in BUCKETS:
            if (now_dt.hour, now_dt.minute) < (h, m):
                bucket_dt = now_dt.replace(hour=h, minute=m, second=0, microsecond=0)
                diff = bucket_dt - now_dt
                minutes = int(diff.total_seconds() // 60)
                if minutes <= 0:
                    status = "ready now"
                elif minutes == 1:
                    status = "in 1 minute"
                else:
                    status = f"in {minutes} minutes"
                return label, f"{bucket_dt.strftime('%I:%M %p').lstrip('0')} ({status})"
        return None, "Session complete"

    def _update_right_rail(
        self,
        label: Optional[str] = None,
        pct_map: Optional[Dict[str, Optional[float]]] = None,
        capture_dt: Optional[datetime] = None,
    ) -> None:
        if not hasattr(self, "right_rail"):
            return
        if pct_map is None or not pct_map:
            pct_map = self._last_pct_map
        if label is None:
            label = self.last_capture_label or self._latest_label_with_data()
        if capture_dt is None:
            capture_dt = self.last_capture_time

        adv = sum(1 for pct in pct_map.values() if isinstance(pct, (int, float)) and pct > 0)
        dec = sum(1 for pct in pct_map.values() if isinstance(pct, (int, float)) and pct < 0)
        flat = len(TICKERS) - adv - dec
        breadth_text = f"Breadth: Adv {adv} · Dec {dec} · Flat {flat}"

        top_ticker: Optional[str] = None
        top_pct: Optional[float] = None
        for tk, pct in pct_map.items():
            if not isinstance(pct, (int, float)):
                continue
            if top_pct is None or abs(pct) > abs(top_pct):
                top_pct = pct
                top_ticker = tk
        if top_ticker is not None and isinstance(top_pct, (int, float)):
            top_text = f"Top mover: {top_ticker} {top_pct:+.2f}%"
        else:
            top_text = "Top mover: --"

        if label is not None:
            if capture_dt is not None:
                time_text = format_et(capture_dt)
                last_capture = f"Last capture: {label} · {time_text}"
            else:
                last_capture = f"Last capture: {label}"
        else:
            last_capture = "Last capture: --"

        next_label, next_phrase = self._next_bucket_context()
        if next_label is not None:
            next_text = f"Next bucket: {next_label} · {next_phrase}"
        else:
            next_text = f"Next bucket: {next_phrase}"

        self.right_rail.update_insights(
            f"Session: {self.session_date.strftime('%A, %b %d, %Y')}",
            last_capture,
            breadth_text,
            top_text,
            next_text,
            self._captured_bucket_count(),
            len(BUCKETS),
            DATA_DIR,
        )

    # ---------- settings / features ----------
    def _load_features(self) -> Dict[str,bool]:
        defaults = {label: False for (label, _) in FEATURE_ITEMS}
        defaults[FEATURE_INSIGHT_RAIL] = True
        if FEATURES_JSON.exists():
            try:
                loaded = json.loads(FEATURES_JSON.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    for key, default in defaults.items():
                        if key not in loaded:
                            loaded[key] = default
                    return loaded
            except Exception:
                pass
        return defaults

    def _save_features(self) -> None:
        try: FEATURES_JSON.write_text(json.dumps(self.features, indent=2), encoding="utf-8")
        except Exception as e: log(f"save-features error: {e}")

    def set_data_folder(self) -> None:
        global DATA_DIR, LOG_PATH, FEATURES_JSON
        d = QFileDialog.getExistingDirectory(self, "Choose data folder for Excel & outputs", str(DATA_DIR))
        if d:
            self.settings["data_dir"] = d
            save_settings(self.settings)
            DATA_DIR = Path(d)
            LOG_PATH = DATA_DIR / "tracker.log"
            FEATURES_JSON = DATA_DIR / "features.json"
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            clear_prior_bucket_cache()
            self.prev_session_bucket = prior_session_bucket(self.session_date)
            self._rerender_table()
            self._update_right_rail()
            self.tray.showMessage(APP_NAME, f"Data folder set to:\n{d}", QSystemTrayIcon.Information, 2500)

    # ---------- app logic ----------
    def initial_sync(self) -> None:
        """
        Perform the first data population when the application starts.  We
        synchronously load the intraday cache and then backfill the grid
        through the current bucket and capture it.  If any exception
        occurs during loading or rendering, it is logged but does not
        crash the UI.
        """
        try:
            _load_intraday_cache()
            self.backfill_to_now()
            self.refresh_now()
            if self.features.get(FEATURE_MORNING_RESUME, False):
                self._maybe_show_morning_resume()
        except Exception as e:
            log(f"initial_sync err: {e}")

    def timer_logic(self) -> None:
        now_dt = now_eastern()
        hm = (now_dt.hour, now_dt.minute)
        if self._last_timer_key == hm:
            return
        self._last_timer_key = hm
        if MARKET_OPEN <= hm <= MARKET_CLOSE:
            self.refresh_now()
        else:
            self.statusBar().showMessage("Outside market hours — next capture at next session open.", 4000)

    def browse_excels(self) -> None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(DATA_DIR))  # type: ignore[attr-defined]
            else:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(DATA_DIR)))
        except Exception as e:
            QMessageBox.warning(self, APP_NAME, f"Could not open: {e}")

    # ---------- capture ----------
    def _bucket_index_now(self, hm: Optional[Tuple[int, int]] = None) -> Optional[int]:
        hour_min = hm or (now_eastern().hour, now_eastern().minute)
        idx = -1
        for i,(_, (h,m)) in enumerate(BUCKETS):
            if hour_min >= (h,m): idx = i
        return idx if idx >= 0 else None

    def backfill_to_now(self) -> None:
        if pd is None: return
        try:
            now_dt = now_eastern()
            hm = (now_dt.hour, now_dt.minute)
            tzinfo = now_dt.tzinfo
            self._ensure_session(now_dt.date())
            idx_now = self._bucket_index_now(hm)
            if idx_now is None: return

            for tk in TICKERS:
                self._ensure_prev_close(tk)

            for b in range(0, idx_now + 1):
                col = 1 + b
                h, m = BUCKETS[b][1]
                label = INDEX_TO_LABEL[b]
                pct_map: Dict[str, Optional[float]] = {}
                for r, tk in enumerate(TICKERS):
                    price = price_at_or_before_bucket(tk, self.session_date, h, m)
                    base = self._base_for(tk, label)
                    _txt, pct, item = self._render_cell(tk, label, price, base)
                    self.table.setItem(r, col, item)
                    self.bucket_prices[tk][label] = price if isinstance(price, (int, float)) else None
                    pct_map[tk] = pct
                self._maybe_export(label)
                capture_dt = datetime(
                    self.session_date.year,
                    self.session_date.month,
                    self.session_date.day,
                    h,
                    m,
                    tzinfo=tzinfo,
                )
                self._post_capture_hooks(label, pct_map, capture_dt)

            last_label = INDEX_TO_LABEL[idx_now]
            self.statusBar().showMessage(f"Backfilled through {last_label} @ {format_et(now_dt)}", 4000)

        except Exception as e:
            log(f"backfill_to_now err: {e}")

    def refresh_now(self) -> None:
        """
        Capture the current bucket.  Before computing new values, we
        refresh the intraday cache so that price_at_or_before_bucket()
        uses up-to-date minute data.  This call is synchronous and may
        take a few seconds, but it eliminates stale data between
        buckets.  Any exceptions are logged and surfaced via a
        message box.
        """
        try:
            _load_intraday_cache()

            now_dt = now_eastern()
            hm = (now_dt.hour, now_dt.minute)
            self._ensure_session(now_dt.date())
            idx = self._bucket_index_now(hm)
            label = INDEX_TO_LABEL.get(idx) if idx is not None else None
            if label is None:
                self.statusBar().showMessage("Pre-market — will capture starting 9:31 AM", 4000)
                return

            label_idx = LABEL_TO_INDEX[label]
            col = 1 + label_idx

            pct_map: Dict[str, Optional[float]] = {}
            for r, tk in enumerate(TICKERS):
                h, m = BUCKETS[label_idx][1]
                price = price_at_or_before_bucket(tk, self.session_date, h, m)
                self._ensure_prev_close(tk)
                base = self._base_for(tk, label)
                _txt, pct, item = self._render_cell(tk, label, price, base)
                self.table.setItem(r, col, item)
                self.bucket_prices[tk][label] = price if isinstance(price, (int, float)) else None
                pct_map[tk] = pct

            self._maybe_export(label)
            self._post_capture_hooks(label, pct_map, now_dt)
            self.statusBar().showMessage(f"Captured {label} @ {format_et(now_dt)}", 4000)

        except Exception:
            log("refresh_now fatal:\n" + traceback.format_exc())
            QMessageBox.warning(self, APP_NAME, "Refresh failed. See tracker.log for details.")

    # ---------- Excel export (Grid + bucket) ----------
    def _snapshot_for_label(self, label: str) -> Tuple[Optional[float], ...]:
        snap: List[Optional[float]] = []
        for tk in TICKERS:
            val = self.bucket_prices.get(tk, {}).get(label)
            snap.append(val if isinstance(val, (int, float)) else None)
        return tuple(snap)

    def _maybe_export(self, label: str, force: bool = False) -> None:
        if pd is None or openpyxl is None:
            return
        snap = self._snapshot_for_label(label)
        if force or self._exported_states.get(label) != snap:
            self._export_excel_grid_and_bucket(label)
            self._exported_states[label] = snap

    def _excel_color_for(self, pct: Optional[float]) -> str:
        if isinstance(pct,(int,float)):
            if pct > 0:  return "22C55E"
            if pct < 0:  return "F87171"
        return "94A3B8"

    def _export_excel_grid_and_bucket(self, label: str) -> None:
        if pd is None or openpyxl is None:
            return
        try:
            day = datetime.now()
            fn = DATA_DIR / day_file_name(day)
            grid_name = "Grid"
            bucket_name = safe_sheet_name(label)

            headers = ["Ticker"] + TIME_COLS
            rows: List[List[str]] = []
            for r, tk in enumerate(TICKERS):
                row = [f"{r+1}. {tk}"]
                for c in range(1, 1 + len(TIME_COLS)):
                    it = self.table.item(r, c)
                    row.append(it.text() if it else "")
                rows.append(row)
            df_grid = pd.DataFrame(rows, columns=headers)

            prices: List[Optional[float]] = []
            pcts: List[Optional[float]] = []
            for tk in TICKERS:
                self._ensure_prev_close(tk)
                px = self.bucket_prices.get(tk, {}).get(label)
                base = self._base_for(tk, label)
                pct_val: Optional[float] = None
                if isinstance(px, (int, float)) and isinstance(base, (int, float)) and base:
                    pct_val = ((px / base) - 1.0) * 100.0
                prices.append(px if isinstance(px, (int, float)) else None)
                pcts.append(pct_val)
            df_bucket = pd.DataFrame({"price": prices, "pct": pcts},
                                     index=[f"{i+1}. {t}" for i,t in enumerate(TICKERS)])

            mode = "a" if fn.exists() else "w"
            with pd.ExcelWriter(fn, engine="openpyxl", mode=mode) as xw:
                try:
                    wb = xw.book  # type: ignore[attr-defined]
                    if grid_name in wb.sheetnames:
                        wb.remove(wb[grid_name])
                except Exception:
                    pass
                df_grid.to_excel(xw, sheet_name=grid_name, index=False)

                try:
                    wb = xw.book  # type: ignore[attr-defined]
                    if bucket_name in wb.sheetnames:
                        wb.remove(wb[bucket_name])
                except Exception:
                    pass
                df_bucket.to_excel(xw, sheet_name=bucket_name)

            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font

            wb = load_workbook(fn)
            ws = wb["Grid"]
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                cell.font = Font(bold=True)
            for r, tk in enumerate(TICKERS, start=2):
                c = ws.cell(row=r, column=1)
                c.font = Font(bold=True, color="000000FF" if tk in ("INTC","WBA") else "00000000")
            for r in range(2, 2 + len(TICKERS)):
                for c in range(2, 2 + len(TIME_COLS) + 0):
                    val = ws.cell(row=r, column=c).value
                    if isinstance(val, str):
                        _, p = parse_price_pct(val)
                        fill = PatternFill("solid", fgColor=self._excel_color_for(p))
                        ws.cell(row=r, column=c).fill = fill

            bws = wb[bucket_name]
            for r in range(2, 2 + len(TICKERS)):
                pval = bws.cell(row=r, column=3).value  # pct column
                try:
                    p = float(pval) if pval not in (None,"") else None
                except Exception:
                    p = None
                fill = PatternFill("solid", fgColor=self._excel_color_for(p))
                bws.cell(row=r, column=2).fill = fill  # price col
                bws.cell(row=r, column=3).fill = fill  # pct col
            wb.save(fn)

        except Exception as e:
            log(f"excel export err: {e}")

    # ---------- Features ----------
    def open_features(self) -> None:
        dlg = QDialog(self); dlg.setWindowTitle("Features"); dlg.setWindowIcon(self._icon())
        v = QVBoxLayout(dlg)
        v.setContentsMargins(18, 18, 18, 18)
        v.setSpacing(12)
        hdr = QLabel("Toggle the extras you want to see. Each switch animates as you enable or disable it so you always know what just changed.")
        hdr.setWordWrap(True)
        hdr.setStyleSheet("color:#cbd5f5; font-size:13px; line-height:150%;")
        v.addWidget(hdr)

        self.boxes: Dict[str, AnimatedToggle] = {}
        for label, tip in FEATURE_ITEMS:
            card = QFrame(dlg)
            card.setFrameShape(QFrame.NoFrame)
            card.setStyleSheet(
                "QFrame { background-color:#0f172a; border:1px solid rgba(148,163,184,0.25); border-radius:10px; }"
            )
            row = QHBoxLayout(card)
            row.setContentsMargins(16, 14, 16, 14)
            row.setSpacing(14)

            copy = QVBoxLayout()
            title = QLabel(label)
            title.setStyleSheet("color:#e2e8f0; font-size:14px; font-weight:600;")
            desc = QLabel(tip)
            desc.setWordWrap(True)
            desc.setStyleSheet("color:#94a3b8; font-size:12px;")
            copy.addWidget(title)
            copy.addWidget(desc)
            copy.addStretch(1)

            toggle = AnimatedToggle(card)
            toggle.setChecked(self.features.get(label, False))
            toggle.setToolTip(tip)
            toggle.setFixedSize(toggle.sizeHint())

            row.addLayout(copy, 1)
            row.addWidget(toggle, 0, Qt.AlignVCenter)

            v.addWidget(card)
            self.boxes[label] = toggle

        v.addStretch(1)

        row = QHBoxLayout(); btnC = QPushButton("Cancel"); btnS = QPushButton("Save"); row.addStretch(1); row.addWidget(btnC); row.addWidget(btnS)
        v.addSpacing(6); v.addLayout(row)

        help_row = QHBoxLayout()
        help_btn = QPushButton("Click this if you are confused about a feature above")
        help_btn.setStyleSheet("QPushButton{padding:8px 10px; font-weight:600;}")
        help_btn.clicked.connect(self._open_feature_explainer)
        help_row.addStretch(1); help_row.addWidget(help_btn)
        v.addSpacing(8); v.addLayout(help_row)

        btnC.clicked.connect(dlg.reject); btnS.clicked.connect(dlg.accept)
        if dlg.exec_() == QDialog.Accepted:
            for k, cb in self.boxes.items(): self.features[k] = cb.isChecked()
            self._save_features()
            self._apply_feature_visibility()
            self._rerender_table()
            self.statusBar().showMessage("Features updated.", 1500)
        else:
            for k, cb in self.boxes.items():
                cb.setChecked(self.features.get(k, False))

        dlg.resize(760, 660)

    def open_replay(self) -> None:
        if not self.features.get(FEATURE_REPLAY, False):
            QMessageBox.information(self, APP_NAME, "Enable Replay Mode in Features first.")
            return
        if openpyxl is None:
            QMessageBox.warning(self, APP_NAME, "Replay Mode requires openpyxl installed.")
            return
        today_file = DATA_DIR / day_file_name(datetime(self.session_date.year, self.session_date.month, self.session_date.day))
        if not today_file.exists():
            today_file = DATA_DIR / day_file_name(datetime.now())
        if not today_file.exists():
            QMessageBox.information(self, APP_NAME, "No workbook for today yet — capture a bucket first.")
            return
        dlg = ReplayDialog(self, today_file)
        dlg.exec_()

    def _open_feature_explainer(self) -> None:
        html = """
        <h2>Feature Explanations</h2>
        <ol>
          <li><b>Mini \"Dow Dashboard\" Pulse</b> — live banner with mood dial, advancers/decliners, concentration, and top mover.</li>
          <li><b>Auto News Ping for Movers</b> — tray alert that pairs the lead mover with its freshest Yahoo headline.</li>
          <li><b>Chart Sparkline View</b> — Unicode sparkline using 1-minute closes so you can see the slope without leaving the grid.</li>
          <li><b>Replay Mode</b> — scrub saved captures like a DVR to replay the session or clip highlights.</li>
          <li><b>Morning Resume Summary</b> — sunrise toast recapping yesterday's leaders, laggards, and index finish.</li>
          <li><b>Historical Echo Mode</b> — on ±2% swings, surface the closest historical analog and what happened next day.</li>
          <li><b>Dow \"Concentration\" Meter</b> — gauge showing how much of the move is driven by the top five components.</li>
          <li><b>Custom Market Sounds</b> — playful chimes whenever a component crosses the percentage threshold you set.</li>
          <li><b>Candle Ghosts</b> — overlay a ghost line at yesterday's close so you instantly know where price is anchored.</li>
          <li><b>Daily Confidence Meter</b> — animated breadth bar tracking advancers out of 30 as the session unfolds.</li>
          <li><b>\"Dow DNA\" Export</b> — weekly CSV drop full of correlations, realized volatility, and stat blocks.</li>
        </ol>
        """
        dlg = QDialog(self); dlg.setWindowTitle("Feature Explainer"); dlg.setWindowIcon(self._icon())
        lay = QVBoxLayout(dlg)
        te = QTextEdit(); te.setReadOnly(True); te.setHtml(html)
        lay.addWidget(te)
        btn = QPushButton("Close"); btn.clicked.connect(dlg.close); lay.addWidget(btn)
        dlg.resize(720, 560)
        dlg.exec_()

    # ---------- guide ----------
    def show_guide(self) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("Where the hell is the thing?")
        dlg.setWindowIcon(self._icon())
        layout = QVBoxLayout(dlg)

        html = """
        <h2>DOW 30 Tracker — Quick Tour</h2>
        <ul>
          <li><b>Toolbar:</b> Refresh · Browse Excels · Auto-Start toggles · view options (Show Times, Show %, Arrows, Stripe Rows) · Features · this guide.</li>
          <li><b>Table:</b> Tickers on the left (1–32); columns are market-hour buckets. Each cell = price vs the previous bucket (9:31 compares to the prior session's 4:00 PM capture, falling back to the latest close when needed).</li>
          <li><b>Status bar:</b> Messages like "Captured 3:00 PM".</li>
          <li><b>System tray:</b> Close hides to tray; launching app again brings this window forward.</li>
          <li><b>Excel:</b> Every capture writes to <code>Sheet__MM_DD_YYYY.xlsx</code> in your data folder (Grid + one sheet per bucket).</li>
        </ul>
        <p>Use "Set Data Folder…" to choose where workbooks are saved. Remembered in <code>settings.json</code>.</p>
        """
        te = QTextEdit(); te.setReadOnly(True); te.setHtml(html)
        layout.addWidget(te)
        btn = QPushButton("Got it"); btn.clicked.connect(dlg.accept); layout.addWidget(btn)
        dlg.resize(720, 560)
        dlg.exec_()

    # ---------- tray / window ----------
    def show_main(self) -> None:
        self.showNormal(); self.activateWindow(); self.raise_()

    def exit_app(self) -> None:
        QApplication.quit()

    def closeEvent(self, e: QCloseEvent) -> None:
        e.ignore(); self.hide()
        self.tray.showMessage(APP_NAME, "Still running in tray. Launch again to bring the window forward.", QSystemTrayIcon.Information, 1800)

    # ---------- autostart ----------
    def enable_autostart(self) -> None:
        if not sys.platform.startswith("win"):
            QMessageBox.information(self, APP_NAME, "Auto-start is supported on Windows only."); return
        try:
            import winreg as wr  # type: ignore
            with wr.OpenKey(wr.HKEY_CURRENT_USER, r"Software\\Microsoft\\Windows\\CurrentVersion\\Run", 0, wr.KEY_SET_VALUE) as k:
                wr.SetValueEx(k, APP_NAME, 0, wr.REG_SZ, exe_path())
            self.tray.showMessage(APP_NAME, "Auto-start enabled.", QSystemTrayIcon.Information, 1500)
        except Exception as e:
            QMessageBox.warning(self, APP_NAME, f"Failed to enable auto-start:\\n{e}")

    def disable_autostart(self) -> None:
        if not sys.platform.startswith("win"):
            QMessageBox.information(self, APP_NAME, "Auto-start is supported on Windows only."); return
        try:
            import winreg as wr  # type: ignore
            with wr.OpenKey(wr.HKEY_CURRENT_USER, r"Software\\Microsoft\\Windows\\CurrentVersion\\Run", 0, wr.KEY_SET_VALUE) as k:
                try: wr.DeleteValue(k, APP_NAME)
                except FileNotFoundError: pass
            self.tray.showMessage(APP_NAME, "Auto-start disabled.", QSystemTrayIcon.Information, 1500)
        except Exception as e:
            QMessageBox.warning(self, APP_NAME, f"Failed to disable auto-start:\\n{e}")

    # ---------- misc ----------
    def _icon(self) -> QIcon:
        for rel in ("assets/dow.ico","assets/dow.png"):
            p = resource_path(rel)
            if os.path.exists(p): return QIcon(p)
        inst = QApplication.instance()
        return inst.style().standardIcon(QStyle.SP_ComputerIcon) if inst else QIcon()

# -------- main --------
def main() -> None:
    ensure_single_instance()
    sock = start_primary_socket()
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME); app.setQuitOnLastWindowClosed(False)
    w = MainWindow()
    socket_listener(sock, on_show=w.show_main)
    w.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"FATAL: {e}\\n{traceback.format_exc()}")
        raise
